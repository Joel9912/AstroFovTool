# utils.py

import os
import csv
from pathlib import Path
from tkinter import messagebox
import openpyxl
from openpyxl.utils import get_column_letter

ARCHIVO_SALIDA = Path(__file__).resolve().parent.parent / "historial" / "objetos guardados.csv"
ARCHIVO_SALIDA.parent.mkdir(parents=True, exist_ok=True)

def guardar_objeto_en_csv(self, nombre, deg_x, deg_y, resultados, marca, tipo, mp):
    if not self.resultados:
        messagebox.showwarning("Sin datos", "Primero debe calcular antes de guardar.")
        return
    filas_guardadas = []
    if os.path.exists(ARCHIVO_SALIDA):
        with open(ARCHIVO_SALIDA, newline="") as f:
            reader = csv.reader(f)
            filas_guardadas = list(reader)

    clave = (nombre, marca, tipo, mp)

    def extraer_clave_bloque(fila1, fila2):
        return (
            fila1[0].strip(),  # nombre
            fila2[0].strip(),  # marca
            fila2[1].strip(),  # tipo
            fila2[2].strip()   # mp
        )

    nuevas_filas_guardadas = []
    i = 0
    while i < len(filas_guardadas):
        if i + 1 < len(filas_guardadas):
            fila1 = filas_guardadas[i]
            fila2 = filas_guardadas[i + 1]
            if len(fila1) > 0 and len(fila2) > 2 and extraer_clave_bloque(fila1, fila2) == clave:
                i += 2
                while i < len(filas_guardadas) and any(filas_guardadas[i]):
                    i += 1
                if i < len(filas_guardadas) and not any(filas_guardadas[i]):
                    i += 1
                continue
        nuevas_filas_guardadas.append(filas_guardadas[i])
        i += 1

    fila_objeto = [nombre, f"{deg_x:.6f}", f"{deg_y:.6f}", "", "", "", "", "", ""]
    fila_sensor = [ marca, tipo, mp, "", "", "", "", "",""]
    nuevas_filas = [fila_objeto, fila_sensor]

    for res in resultados:
        nuevas_filas.append([
            "", "", "",
            f"{res['f']}mm",
            f"{res['px_x']:.1f}",
            f"{res['px_y']:.1f}",
            f"{res['total']:.0f}",
            f"{res['porcentaje']:.3f}",
            "Sí" if res["fotografiable"] else "No"
        ])

    nuevas_filas.append([])
    nuevas_filas.append([])
    nuevas_filas.append([])

    with open(str(ARCHIVO_SALIDA), "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(nuevas_filas_guardadas + nuevas_filas)

    return True

def leer_historial():
    if not os.path.exists(ARCHIVO_SALIDA):
        return []
    with open(ARCHIVO_SALIDA, newline="") as f:
        reader = csv.reader(f)
        return list(reader)

def exportar_excel_desde_csv():
    if not os.path.exists(ARCHIVO_SALIDA):
        messagebox.showerror("Error", "No hay datos para exportar.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial DSO"

    headers = [
        "Nombre/Marca", "° Ancho/Tipo", "° Alto/MP", "Focal", "Px X", "Px Y",
         "Total de Pixeles", "% del Sensor", "¿Fotografiable?"
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    with open(ARCHIVO_SALIDA, newline="") as f:
        reader = csv.reader(f)
        datos = list(reader)

    fila_excel = 2
    i = 0
    while i < len(datos):
        fila1 = datos[i] if i < len(datos) else []
        fila2 = datos[i + 1] if i + 1 < len(datos) else []
        i += 2
        if not fila1 or not fila2:
            continue

        nombre = fila1[0] if len(fila1) > 0 else ""
        deg_x = fila1[1] if len(fila1) > 1 else ""
        deg_y = fila1[2] if len(fila1) > 2 else ""
        marca = fila2[0] if len(fila2) > 0 else ""
        tipo = fila2[1] if len(fila2) > 1 else ""
        mp = fila2[2] if len(fila2) > 2 else ""

        index_fila = 0
        while i < len(datos) and any(datos[i]):
            fila = datos[i]
            focal = fila[3] if len(fila) > 3 else ""
            px_x = fila[4] if len(fila) > 4 else ""
            px_y = fila[5] if len(fila) > 5 else ""
            total = fila[6] if len(fila) > 6 else ""
            pct = fila[7] if len(fila) > 7 else ""
            fotogra = fila[8] if len(fila) > 8 else ""

            if index_fila == 0:
                fila_valores = [nombre, deg_x, deg_y, focal, px_x, px_y, total, pct, fotogra]
            elif index_fila == 1:
                fila_valores = [marca, tipo, mp, focal, px_x, px_y, total, pct, fotogra]
            else:
                fila_valores = ["", "", "", focal, px_x, px_y, total, pct, fotogra]

            for col, valor in enumerate(fila_valores, start=1):
                ws.cell(row=fila_excel, column=col, value=valor)
            fila_excel += 1
            i += 1
            index_fila += 1

        fila_excel += 3  # Saltar 3 filas para separación

        if i < len(datos) and not any(datos[i]):
            i += 1

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    ruta_excel = str(ARCHIVO_SALIDA).replace(".csv", ".xlsx")
    wb.save(ruta_excel)
    messagebox.showinfo("Éxito", f"Historial exportado a:\n{ruta_excel}")

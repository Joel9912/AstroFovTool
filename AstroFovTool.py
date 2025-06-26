import tkinter as tk
from tkinter import ttk
import math
import os
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import messagebox
from pathlib import Path


# ========================
# CONFIGURACIÓN DEL SENSOR
# ========================
SENSOR_WIDTH_MM = 23.5
SENSOR_HEIGHT_MM = 15.6
RES_X = 6016
RES_Y = 4016
PIXEL_SIZE_X = SENSOR_WIDTH_MM / RES_X
PIXEL_SIZE_Y = SENSOR_HEIGHT_MM / RES_Y

# ======================
# PARÁMETROS DE CÁLCULO
# ======================
FOCALES_ESTANDAR = [18, 24, 35, 50, 70, 85, 100, 150, 200, 300, 500, 650, 750, 900, 1000, 1250, 1500, 1800]
UMBRAL_PIXELES = 15000
CANVAS_WIDTH = 300
CANVAS_HEIGHT = 200

# Ruta personalizada
ARCHIVO_SALIDA = Path("D:/Usuario/Escritorio/Astro/AstroFovTool/objetos_guardados.csv")
ARCHIVO_SALIDA.parent.mkdir(parents=True, exist_ok=True)

# =======================
# FUNCIONES AUXILIARES
# =======================
def dms_a_grados(d, m, s):
    try:
        d_val = float(d) if d.strip() != "" else 0.0
        m_val = float(m) if m.strip() != "" else 0.0
        s_val = float(s) if s.strip() != "" else 0.0
        return abs(d_val) + m_val / 60 + s_val / 3600
    except:
        return None

def calcular_fov_en_px(obj_deg, focal):
    fov_sensor_x = (SENSOR_WIDTH_MM / focal) * (180 / math.pi)
    fov_sensor_y = (SENSOR_HEIGHT_MM / focal) * (180 / math.pi)
    px_por_grado_x = RES_X / fov_sensor_x
    px_por_grado_y = RES_Y / fov_sensor_y
    pixeles_x = obj_deg[0] * px_por_grado_x
    pixeles_y = obj_deg[1] * px_por_grado_y
    total_pixeles = pixeles_x * pixeles_y
    porcentaje = (total_pixeles / (RES_X * RES_Y)) * 100
    fotografiable = total_pixeles >= UMBRAL_PIXELES
    return pixeles_x, pixeles_y, total_pixeles, porcentaje, fotografiable

# =========================
# CLASE PRINCIPAL DE APP
# =========================
class AstroFovToolApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("AstroFovTool")
        self.geometry("800x600")
        self.modo = tk.StringVar(value="tabla")
        self.crear_widgets()
        self.actualizar_estado_guardado()

    def crear_widgets(self):
        frame_izq = ttk.Frame(self)
        frame_der = ttk.Frame(self, width=500)
        frame_izq.pack(side="left", fill="y", padx=10, pady=10)
        frame_der.pack(side="left", fill="both", padx=10, pady=10)

        frame_sensor = ttk.LabelFrame(frame_izq, text="Datos del sensor Nikon D3300")
        frame_sensor.pack(fill="x", pady=5)
        ttk.Label(frame_sensor, text=f"Tamaño sensor: {SENSOR_WIDTH_MM} mm x {SENSOR_HEIGHT_MM} mm").pack(anchor="w", padx=10, pady=2)
        ttk.Label(frame_sensor, text=f"Resolución sensor: {RES_X} px x {RES_Y} px").pack(anchor="w", padx=10, pady=2)
        ttk.Label(frame_sensor, text=f"Tamaño pixel: {PIXEL_SIZE_X * 1000:.3f} µm x {PIXEL_SIZE_Y * 1000:.3f} µm").pack(anchor="w", padx=10, pady=2)

        frame_objeto = ttk.LabelFrame(frame_izq, text="Tamaño del objeto en cielo (DSO)")
        frame_objeto.pack(fill="x", pady=5)
        self.entries = {}

        # Etiqueta para el nombre del objeto
        ttk.Label(frame_objeto, text="Nombre del objeto:").grid(row=2, column=0, sticky="w", padx=5, pady=2)
        self.nombre_objeto = tk.Entry(frame_objeto)
        self.nombre_objeto.grid(row=2, column=1, columnspan=6, sticky="we")

        # Entradas para ancho y alto con unidades separadas
        for i, label_text in enumerate(["Ancho:", "Alto:"]):
            ttk.Label(frame_objeto, text=label_text).grid(row=i, column=0, sticky="w", padx=5, pady=2)
            self.entries[i] = []

            for j, simbolo in enumerate(["°", "'", "\""]):
                entry = tk.Entry(frame_objeto, width=4)
                entry.grid(row=i, column=1 + j * 2, padx=1)
                self.entries[i].append(entry)
                entry.bind("<Left>", lambda e, i=i, j=j: self.mover_entre_entradas(e, i, j))
                entry.bind("<Right>", lambda e, i=i, j=j: self.mover_entre_entradas(e, i, j))
                ttk.Label(frame_objeto, text=simbolo).grid(row=i, column=2 + j * 2, sticky="w", padx=(0, 5))

        frame_focal = ttk.LabelFrame(frame_izq, text="Configuración")
        frame_focal.pack(fill="x", pady=5)
        ttk.Radiobutton(frame_focal, text="Modo tabla (focales estándar)", variable=self.modo, value="tabla").pack(anchor="w")
        ttk.Radiobutton(frame_focal, text="Modo focal fija", variable=self.modo, value="fija").pack(anchor="w")

        self.focal_entry = tk.Entry(frame_focal)
        self.focal_entry.pack(anchor="w", padx=20)
        self.focal_entry.insert(0, "50")

        ttk.Button(frame_izq, text="Calcular", command=self.mostrar_resultados).pack(pady=10)
        self.boton_guardar = ttk.Button(frame_izq, text="Guardar objeto", command=self.guardar_objeto)
        self.boton_guardar.pack(pady=5)
        ttk.Button(frame_izq, text="Ver historial", command=self.ver_historial).pack(pady=5)

        self.tree = ttk.Treeview(frame_der, columns=("Focal", "Px X", "Px Y", "Total", "%"), show="headings")
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
        # Ajustar anchos y centrar texto
        self.tree.column("Focal", width=100, anchor="center")
        self.tree.column("Px X", width=100, anchor="center")
        self.tree.column("Px Y", width=100, anchor="center")
        self.tree.column("Total", width=100, anchor="center")
        self.tree.column("%", width=100, anchor="center")
        self.tree.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(frame_der, width=CANVAS_WIDTH, height=CANVAS_HEIGHT, bg="white")
        self.canvas.pack(pady=10)

        self.tree.bind("<<TreeviewSelect>>", self.on_row_select)

        self.tree.tag_configure("verde", background="lightgreen")
        self.tree.tag_configure("rojo", background="lightcoral")

    def mover_entre_entradas(self, event, i, j):
        if event.keysym == "Right":
            if j < 2:
                self.entries[i][j + 1].focus()
            elif i == 0:
                self.entries[1][0].focus()
            else:
                self.nombre_objeto.focus()
        elif event.keysym == "Left":
            if j > 0:
                self.entries[i][j - 1].focus()
            elif i == 1:
                self.entries[0][2].focus()

    def actualizar_estado_guardado(self):
        if self.modo.get() == "tabla":
            self.boton_guardar.state(["!disabled"])
        else:
            self.boton_guardar.state(["disabled"])

        self.modo.trace_add("write", lambda *args: self.actualizar_estado_guardado())

    def mostrar_resultados(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.canvas.delete("all")

        deg_x = dms_a_grados(*[e.get() for e in self.entries[0]])
        deg_y = dms_a_grados(*[e.get() for e in self.entries[1]])
        if deg_x is None or deg_y is None:
            messagebox.showwarning("Entrada inválida", "Por favor, ingrese valores válidos para ancho y alto.")
            return

        self.resultados = []
        focales = FOCALES_ESTANDAR if self.modo.get() == "tabla" else [float(self.focal_entry.get())]
        for f in focales:
            px_x, px_y, total, pct, fot = calcular_fov_en_px((deg_x, deg_y), f)
            self.resultados.append({
                "f": f, "px_x": px_x, "px_y": px_y,
                "total": total, "porcentaje": pct, "fotografiable": fot
            })

        for res in self.resultados:
            tag = "verde" if res["fotografiable"] else "rojo"
            self.tree.insert("", "end",
                values=(f"{res['f']}mm", f"{res['px_x']:.1f}", f"{res['px_y']:.1f}",
                        f"{res['total']:.0f}", f"{res['porcentaje']:.3f} %"),
                tags=(tag,))

        if self.modo.get() == "fija":
            self.dibujar_canvas(self.resultados[0])

    def on_row_select(self, event):
        if self.modo.get() != "tabla":
            return
        selected = self.tree.selection()
        if selected:
            idx = self.tree.index(selected[0])
            self.dibujar_canvas(self.resultados[idx])

    def dibujar_canvas(self, res):
        self.canvas.delete("all")
        escala_x = CANVAS_WIDTH / RES_X
        escala_y = CANVAS_HEIGHT / RES_Y
        w = res["px_x"] * escala_x
        h = res["px_y"] * escala_y
        x0 = (CANVAS_WIDTH - w) / 2
        y0 = (CANVAS_HEIGHT - h) / 2
        x1 = x0 + w
        y1 = y0 + h
        color = "green" if res["fotografiable"] else "red"
        self.canvas.create_rectangle(0, 0, CANVAS_WIDTH, CANVAS_HEIGHT, outline="gray")
        self.canvas.create_rectangle(x0, y0, x1, y1, outline=color, width=2)
        self.canvas.create_text(CANVAS_WIDTH/2, CANVAS_HEIGHT + 10, anchor="n", text=f"{res['porcentaje']:.3f}% del sensor", fill=color, font=("Arial", 10))

    def guardar_objeto(self):
        if self.modo.get() != "tabla":
            return

        nombre = self.nombre_objeto.get().strip()
        if not nombre or not self.resultados:
            return

        deg_x = dms_a_grados(*[e.get() for e in self.entries[0]])
        deg_y = dms_a_grados(*[e.get() for e in self.entries[1]])

        # Leer todo el archivo actual si existe
        filas_guardadas = []
        if os.path.exists(ARCHIVO_SALIDA):
            with open(ARCHIVO_SALIDA, newline="") as f:
                reader = csv.reader(f)
                filas_guardadas = list(reader)

        # Verificar si el nombre ya existe
        nombre_existente = any(fila and fila[0].strip() == nombre for fila in filas_guardadas if fila)

        if nombre_existente:
            sobrescribir = messagebox.askyesno("Objeto ya guardado",
                                               f"Ya existe un objeto llamado '{nombre}'.\n¿Desea sobrescribirlo?")
            if not sobrescribir:
                return
            # Eliminar entradas anteriores con ese nombre y la línea en blanco siguiente (si la hay)
            nuevas_filas_guardadas = []
            i = 0
            while i < len(filas_guardadas):
                fila = filas_guardadas[i]
                if fila and fila[0].strip() == nombre:
                    # Saltar este bloque
                    i += 1
                    while i < len(filas_guardadas) and filas_guardadas[i] and filas_guardadas[i][0].strip() != "":
                        i += 1
                    # También saltar la línea en blanco siguiente si existe
                    if i < len(filas_guardadas) and (
                            not filas_guardadas[i] or all(c.strip() == "" for c in filas_guardadas[i])):
                        i += 1
                else:
                    nuevas_filas_guardadas.append(fila)
                    i += 1

            filas_guardadas = nuevas_filas_guardadas

        # Agregar nuevas filas
        nuevas_filas = []
        for res in self.resultados:
            nuevas_filas.append([
                nombre,
                f"{deg_x:.6f}",
                f"{deg_y:.6f}",
                f"{res['f']}mm",
                f"{res['px_x']:.1f}",
                f"{res['px_y']:.1f}",
                f"{res['total']:.0f}",
                f"{res['porcentaje']:.3f}",
                "Sí" if res['fotografiable'] else "No"
            ])
        nuevas_filas.append([])  # Línea en blanco al final del bloque

        # Escribir archivo actualizado
        with open(str(ARCHIVO_SALIDA), "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(filas_guardadas + nuevas_filas)

    def ver_historial(self):
        if not os.path.exists(ARCHIVO_SALIDA):
            messagebox.showinfo("Historial vacío", "No se han guardado objetos aún.")
            return

        popup = tk.Toplevel(self)
        popup.title("Historial de objetos guardados")
        popup.geometry("820x400")

        tree_hist = ttk.Treeview(popup, columns=(
        "Nombre", "Ancho", "Alto", "Focal", "Px X", "Px Y", "Total", "%", "¿Fotografiable?"), show="headings")
        for col in tree_hist["columns"]:
            tree_hist.heading(col, text=col)
            tree_hist.column(col, width=80, anchor="center")
        tree_hist.pack(fill="both", expand=True, padx=10, pady=10)

        with open(ARCHIVO_SALIDA, newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                tree_hist.insert("", "end", values=row)

        ttk.Button(popup, text="Exportar a Excel", command=self.exportar_excel).pack(pady=10)

    def exportar_excel(self):
        if not os.path.exists(ARCHIVO_SALIDA):
            messagebox.showerror("Error", "No hay datos para exportar.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial DSO"

        headers = ["Nombre", "Ancho (°)", "Alto (°)", "Focal", "Px X", "Px Y", "Total px", "% Sensor",
                   "¿Fotografiable?"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        with open(ARCHIVO_SALIDA, newline="") as f:
            reader = csv.reader(f)
            for row_idx, fila in enumerate(reader, start=2):
                for col_idx, valor in enumerate(fila, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=valor)

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16

        ruta_excel = str(ARCHIVO_SALIDA).replace(".csv", ".xlsx")
        wb.save(ruta_excel)

        messagebox.showinfo("Éxito", f"Historial exportado a:\n{ruta_excel}")


if __name__ == "__main__":
    app = AstroFovToolApp()
    app.mainloop()
